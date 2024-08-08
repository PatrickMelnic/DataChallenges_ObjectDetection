import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from torch.utils.data import DataLoader
import torch
import numpy as np
os.environ["CUDA_VISIBLE_DEVICES"] = "2"

# Customizable parameters
MODEL_PATH = "vit/checkpoint-8800"  # Path to the trained model
DATASET_PATH = "datasets/multilabel"  # Path to the test dataset
CLASS_MAPPING_FILE = "class_index_mapping.json"  # Path to the class index mapping file
EXCEL_OUTPUT_PATH = "test_results.xlsx"  # Path to save the Excel file
BATCH_SIZE = 1  # Batch size for DataLoader
THRESHOLD = 0.5  # Threshold for predictions

# Set device
device = "cuda" if torch.cuda.is_available() else "cpu"

# Load the model
from transformers import ViTForImageClassification
model = ViTForImageClassification.from_pretrained(MODEL_PATH, ignore_mismatched_sizes=True)
model.to(device)

# Load the dataset
from datasets import load_from_disk
test_dataset = load_from_disk(DATASET_PATH)["test"]

# Load the class index mapping
import json
with open(CLASS_MAPPING_FILE, 'r') as f:
    label_map = json.load(f)

# Create a DataLoader for the test dataset
def collate_fn(batch):
    return {
        "pixel_values": torch.stack([x["pixel_values"] for x in batch]),
        "labels": torch.tensor([x["labels"] for x in batch]),
        "image_path": [x["image_path"] for x in batch],
    }

test_dataloader = DataLoader(test_dataset, batch_size=BATCH_SIZE, collate_fn=collate_fn)

# Function to get predictions
def get_predictions(model, dataloader, threshold=0.5):
    model.eval()
    predictions, true_labels, image_paths = [], [], []
    for batch in dataloader:
        with torch.no_grad():
            outputs = model(pixel_values=batch["pixel_values"].to(device))
            logits = outputs.logits
            preds = (logits.sigmoid() > threshold).int().cpu().numpy()
            predictions.append(preds[0])
            true_labels.append(batch["labels"].cpu().numpy()[0])
            image_paths.append(batch["image_path"][0])
    return predictions, true_labels, image_paths

# Get predictions, true labels, and image paths
predictions, true_labels, image_paths = get_predictions(model, test_dataloader)

# Assuming `label_map` is a dictionary mapping label indices to label names
label_map = {i: f'label_{i}' for i in range(len(true_labels[0]))}

# Calculate the percentage of correct labels predicted
def calculate_percentage(true_labels, predicted_labels):
    correct = sum(1 for t, p in zip(true_labels, predicted_labels) if t == p and t == 1)
    total = sum(true_labels)
    if total == 0:
        return 0
    return correct / total

# Store predictions and true labels in a DataFrame
data = []
for img_path, true_lbls, pred_lbls in zip(image_paths, true_labels, predictions):
    percentage = calculate_percentage(true_lbls, pred_lbls)
    data.append({
        "Image Path": img_path,
        "True Labels": [label_map[i] for i, lbl in enumerate(true_lbls) if lbl],
        "Predicted Labels": [label_map[i] for i, lbl in enumerate(pred_lbls) if lbl],
        "Percentage Correct": percentage
    })

df = pd.DataFrame(data)

# Function to get the color based on the percentage
def get_color(percentage):
    if percentage == 0:
        return "FF0000"  # Red
    elif percentage <= 0.25:
        return "FFA500"  # Orange
    elif percentage <= 0.50:
        return "FFFF00"  # Yellow
    elif percentage <= 0.75:
        return "90EE90"  # Light Green
    elif percentage < 1:
        return "008000"  # Green
    else:
        return "0000FF"  # Blue

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Test Results"

# Add headers
ws.append(["Image Path", "True Labels", "Predicted Labels", "Percentage Correct"])

# Add data rows with conditional formatting
for index, row in df.iterrows():
    color = get_color(row["Percentage Correct"])
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    ws.append([row["Image Path"], ", ".join(row["True Labels"]), ", ".join(row["Predicted Labels"]), row["Percentage Correct"]])
    
    for col in range(1, 5):  # Assuming 4 columns
        ws.cell(row=index+2, column=col).fill = fill

# Save the workbook
wb.save(EXCEL_OUTPUT_PATH)
