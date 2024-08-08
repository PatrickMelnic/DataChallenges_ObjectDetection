import torch
import numpy as np
import pandas as pd
from torchvision import transforms
from torch.utils.data import Dataset, DataLoader
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import json
from PIL import Image
from tqdm import tqdm
import os
import random
import time


# Set the GPU device
os.environ["CUDA_VISIBLE_DEVICES"] = "3"
# Load class index mapping
print(f"{datetime.now()}: Loading class index mapping...")
with open('class_index_mapping.json', 'r') as f:
    class_index_mapping = json.load(f)

# Load the model
print(f"{datetime.now()}: Loading model...")
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
model_path = 'models/swin/'  # Adjust as needed
try:
    model = torch.load(model_path)  # This assumes model is saved with torch.save(model)
except Exception as e:
    print(f"Error loading full model: {e}")
    # If using Hugging Face's save format
    from transformers import SwinForImageClassification
    model = SwinForImageClassification.from_pretrained(model_path)
model.to(device)
model.eval()

# Load the dataset
print(f"{datetime.now()}: Loading dataset...")
from datasets import load_from_disk
DATASET_PATH = 'datasets/multilabel'
dataset_dict = load_from_disk(DATASET_PATH)
test_dataset = dataset_dict["test"]


# Dataset class definition
class MultiLabelTestDataset(Dataset):
    def __init__(self, dataset, transform=None):
        self.dataset = dataset
        self.transform = transform

    def __len__(self):
        return len(self.dataset)

    def __getitem__(self, idx):
        sample = self.dataset[idx]
        image = sample['image']  # Using the loaded PIL image object
        if self.transform:
            image = self.transform(image)
        return image, sample['encoded_label'], sample['image_path'], sample['absolute_path']

# Data loading
transform = transforms.Compose([transforms.ToTensor()])
test_dataset = MultiLabelTestDataset(test_dataset, transform=transform)
test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False)



# Predict and collect data
"""
print(f"{datetime.now()}: Starting prediction and data collection...")
data_dict = {'image': [], 'absolute_path': [], 'true_labels': [], 'predicted_labels': []}
for images, labels, paths, abs_paths in tqdm(test_loader):
    images = images.to(device)
    outputs = model(images).logits
    predictions = torch.sigmoid(outputs).cpu().detach().numpy() > 0.5
    data_dict['image'].extend(paths)
    data_dict['absolute_path'].extend(abs_paths)
    data_dict['true_labels'].append(labels.tolist())
    data_dict['predicted_labels'].extend(predictions.tolist())
"""

# Create a reverse mapping from index to label name
index_to_label = {v: k for k, v in class_index_mapping.items()}

# Predict and collect data

print(f"{datetime.now()}: Starting prediction and data collection...")
result_dict = {'image': [], 'absolute_path': [], 'true_labels': [], 'predicted_labels': []}
probability_threshold = 0.5  # Set a threshold for predictions

for images, labels, paths, abs_paths in tqdm(test_loader):
    images = images.to(device)
    outputs = model(images).logits
    probabilities = torch.sigmoid(outputs).cpu().detach().numpy()  # Get probabilities

    # Convert probabilities to human-readable labels, filtered by the threshold
    human_readable_predictions = []
    for prediction in probabilities:
        label_names = [index_to_label[i] for i, probability in enumerate(prediction) if probability > probability_threshold]
        human_readable_predictions.append(label_names)
    
    result_dict['image'].extend(paths)
    result_dict['absolute_path'].extend(abs_paths)
    result_dict['true_labels'].append(labels)  # Assuming labels are already in a readable format
    result_dict['predicted_labels'].append(human_readable_predictions)

# Print first few entries to verify the content of result_dict
print("First 5 entries of the results:")
for i in range(min(1, len(result_dict['image']))):
    print(f"Entry {i+1}:")
    print(f"Image Path: {result_dict['image'][i]}")
    time.sleep(1)
    print(f"Absolute Path: {result_dict['absolute_path'][i]}")
    time.sleep(1)
    print(f"True Labels: {result_dict['true_labels'][i]}")
    time.sleep(1)
    print(f"Predicted Labels: {result_dict['predicted_labels'][i]}")
    time.sleep(1)
    print("\n")
    
# Save data to Excel with conditional formatting
print(f"{datetime.now()}: Saving data to Excel...")
wb = Workbook()
ws = wb.active
headers = ['Image', 'Absolute Path', 'True Labels', 'Predicted Labels']
ws.append(headers)
for i in range(len(data_dict['image'])):
    row = [
        data_dict['image'][i],
        data_dict['absolute_path'][i],
        ', '.join([class_index_mapping[str(idx)] for idx in data_dict['true_labels'][i] if idx == 1]),
        ', '.join([class_index_mapping[str(idx)] for idx in data_dict['predicted_labels'][i] if idx == 1])
    ]
    ws.append(row)

for i in range(2, ws.max_row + 1):
    true_labels = set(data_dict['true_labels'][i-2])
    predicted_labels = set(data_dict['predicted_labels'][i-2])
    intersection = true_labels & predicted_labels
    union = true_labels | predicted_labels
    similarity_ratio = len(intersection) / len(union) if union else 0
    if similarity_ratio == 1.0:
        fill = PatternFill(start_color='00C851', end_color='00C851', fill_type='solid')
    elif similarity_ratio == 0:
        fill = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
    else:
        green = int(204 * similarity_ratio + 68)
        red = 255 - green
        color_code = f'{red:02X}{green:02X}44'
        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
    ws['D' + str(i)].fill = fill
wb.save('test_results.xlsx')

# Generate heatmaps
print(f"{datetime.now()}: Generating heatmaps...")
num_labels = len(class_index_mapping)
label_names = list(class_index_mapping.values())
true_label_matrix = np.zeros((num_labels, num_labels))
predicted_label_matrix = np.zeros((num_labels, num_labels))
for labels in data_dict['true_labels']:
    indices = [int(label) for label in labels]
    for i in indices:
        for j in indices:
            true_label_matrix[i][j] += 1
for labels in data_dict['predicted_labels']:
    indices = [int(label) for label in labels]
    for i in indices:
        for j in indices:
            predicted_label_matrix[i][j] += 1
true_label_matrix /= true_label_matrix.max()
predicted_label_matrix /= predicted_label_matrix.max()
fig, ax = plt.subplots(1, 2, figsize=(18, 8))
sns.heatmap(true_label_matrix, annot=True, fmt=".2f", ax=ax[0], cmap='viridis', xticklabels=label_names, yticklabels=label_names)
ax[0].set_title('True Labels Correlation')
sns.heatmap(predicted_label_matrix, annot=True, fmt=".2f", ax=ax[1], cmap='viridis', xticklabels=label_names, yticklabels=label_names)
ax[1].set_title('Predicted Labels Correlation')
plt.show()
print(f"{datetime.now()}: Completed all tasks.")
